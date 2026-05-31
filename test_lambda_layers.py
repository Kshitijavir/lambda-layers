import json
import sys
import io
import psycopg2
import pandas as pd
import openpyxl
import requests
import xlsxwriter


def lambda_handler(event, context):
    results = []

    # =============================================
    # 1. Test psycopg2
    # =============================================
    results.append({
        "layer": "psycopg2",
        "status": "✅ OK",
        "version": psycopg2.__version__,
        "info": "PostgreSQL adapter loaded successfully"
    })
    print("[LAYER TEST] psycopg2 ✅ - Version:", psycopg2.__version__)

    # =============================================
    # 2. Test pandas
    # =============================================
    df = pd.DataFrame({"col": [1, 2, 3]})
    results.append({
        "layer": "pandas",
        "status": "✅ OK",
        "version": pd.__version__,
        "info": f"DataFrame test passed: {df.shape[0]} rows x {df.shape[1]} cols"
    })
    print("[LAYER TEST] pandas ✅ - Version:", pd.__version__)

    # =============================================
    # 3. Test openpyxl
    # =============================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    results.append({
        "layer": "openpyxl",
        "status": "✅ OK",
        "version": openpyxl.__version__,
        "info": "Workbook created and cell written successfully"
    })
    print("[LAYER TEST] openpyxl ✅ - Version:", openpyxl.__version__)

    # =============================================
    # 4. Test requests
    # =============================================
    resp = requests.get("https://jsonplaceholder.typicode.com/todos/1", timeout=5)
    results.append({
        "layer": "requests",
        "status": "✅ OK",
        "version": requests.__version__,
        "info": f"HTTP GET successful - Status: {resp.status_code}"
    })
    print("[LAYER TEST] requests ✅ - Version:", requests.__version__)

    # =============================================
    # 5. Test XlsxWriter
    # =============================================
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, "Test")
    workbook.close()
    results.append({
        "layer": "XlsxWriter",
        "status": "✅ OK",
        "version": xlsxwriter.__version__,
        "info": "Excel file created in memory successfully"
    })
    print("[LAYER TEST] XlsxWriter ✅ - Version:", xlsxwriter.__version__)

    # =============================================
    # Summary
    # =============================================
    passed = sum(1 for r in results if r["status"] == "✅ OK")
    failed = sum(1 for r in results if r["status"] == "❌ FAIL")
    total = len(results)

    print("\n" + "=" * 50)
    print(f"LAYER TEST SUMMARY: {passed}/{total} passed, {failed} failed")
    print("=" * 50)

    return {
        "statusCode": 200,
        "body": json.dumps({
            "python_version": sys.version,
            "total_layers": total,
            "passed": passed,
            "failed": failed,
            "results": results
        }, indent=2)
    }
